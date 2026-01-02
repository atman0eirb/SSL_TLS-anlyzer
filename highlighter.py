import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl import load_workbook

# Load CSV
df = pd.read_csv("ssl_output/summary.csv")

# Export to Excel
excel_file = "ssl_output/summary_highlighted.xlsx"
df.to_excel(excel_file, index=False)

# Open workbook
wb = load_workbook(excel_file)
ws = wb.active

# Define colors
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")    # Insecure (Oui)
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Secure (Non)

# Columns to treat "Oui" as secure (green)
tls_good = ["TLS1.2", "TLS1.3"]
# Apply highlighting
for row in ws.iter_rows(min_row=2):  # Skip header
    for cell in row:
        header = ws.cell(row=1, column=cell.column).value  # get column header
        value = str(cell.value).strip().lower()
        
        if value == "oui":
            if header in tls_good:
                cell.fill = green_fill  # TLS 1.2/1.3 "Oui" is good → green
            else:
                cell.fill = red_fill    # other "Oui" → red
        elif value == "non":
            cell.fill = green_fill      # Non → green

# Save
wb.save(excel_file)
print(f"[+] Excel created with highlights: {excel_file}")
